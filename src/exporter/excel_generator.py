#!/usr/bin/env python3
"""
Генератор Excel отчетов с продвинутым форматированием
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path

from ..utils.logger import setup_logger
from ..utils.error_handler import retry_on_failure
from .data_formatter import DataFormatter
from .template_manager import TemplateManager


class ExcelGenerator:
    """Продвинутый генератор Excel отчетов"""
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        self.logger = setup_logger("excel_generator")
        self.data_formatter = DataFormatter()
        self.template_manager = TemplateManager()
        
        self.export_settings = {
            'auto_adjust_columns': True,
            'include_images': False,
            'apply_formatting': True,
            'multiple_sheets': True,
            'compression': True,
            'new_structure': True  # НОВАЯ НАСТРОЙКА: использовать новую структуру
        }
        
        if 'export' in self.config:
            self.export_settings.update(self.config['export'])
    
    @retry_on_failure(max_retries=2)
    def generate_comprehensive_report(self, 
                                   products: List[Dict],
                                   price_comparison: Optional[Dict] = None,
                                   trend_analysis: Optional[Dict] = None,
                                   validation_results: Optional[Dict] = None,
                                   alerts: Optional[Dict] = None,
                                   site_url: str = "") -> str:
        """
        Генерация комплексного Excel отчета
        
        Args:
            products: Список товаров
            price_comparison: Результаты сравнения цен
            trend_analysis: Результаты анализа трендов
            validation_results: Результаты валидации
            alerts: Уведомления системы
            site_url: URL целевого сайта
            
        Returns:
            str: Путь к созданному файлу
        """
        self.logger.info("💾 Создание комплексного Excel отчета")
        
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"comprehensive_report_{timestamp}.xlsx"
            filepath = Path("data/output/excel_exports") / filename
            
            # Создание Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # НОВАЯ СТРУКТУРА: Основной лист с товарами в новой структуре
                if self.export_settings['new_structure']:
                    self._add_products_sheet_new_structure(writer, products, "Товары")
                else:
                    self._add_products_sheet(writer, products, "Товары")
                
                # Аналитические листы
                if price_comparison and price_comparison.get('success'):
                    self._add_price_comparison_sheet(writer, price_comparison, "Сравнение цен")
                
                if trend_analysis and trend_analysis.get('success'):
                    self._add_trend_analysis_sheet(writer, trend_analysis, "Анализ трендов")
                
                if validation_results and validation_results.get('success'):
                    self._add_validation_sheet(writer, validation_results, "Валидация данных")
                
                if alerts and alerts.get('summary', {}).get('total_alerts', 0) > 0:
                    self._add_alerts_sheet(writer, alerts, "Уведомления")
                
                # Сводный лист
                self._add_summary_sheet(writer, products, price_comparison, 
                                      trend_analysis, validation_results, alerts, site_url)
            
            # Применение форматирования
            if self.export_settings['apply_formatting']:
                self._apply_advanced_formatting(filepath)
            
            self.logger.info(f"✅ Excel отчет создан: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"❌ Ошибка создания Excel отчета: {e}")
            raise

    def _add_products_sheet_new_structure(self, writer: pd.ExcelWriter, products: List[Dict], sheet_name: str):
        """
        НОВЫЙ МЕТОД: Добавление листа с товарами в новой структуре
        
        Структура: Категория | Артикул | Фотографии | Описание | [все характеристики] | Цена
        """
        if not products:
            return
        
        self.logger.info("📊 Формирование новой структуры отчета")
        
        # Собираем все возможные характеристики из всех товаров
        all_characteristics = self._collect_all_characteristics(products)
        self.logger.info(f"📋 Найдено характеристик: {len(all_characteristics)}")
        
        # Формируем данные для DataFrame
        rows = []
        for product in products:
            row = {
                'Категория': product.get('category', ''),
                'Артикул': product.get('article', ''),
                'Фотографии': self._format_images_field(product.get('processed_images', {})),
                'Описание': product.get('description', ''),
                'Цена': product.get('price', '')
            }
            
            # Добавляем все характеристики
            characteristics = product.get('characteristics', {})
            for char_name in all_characteristics:
                row[char_name] = characteristics.get(char_name, '')
            
            rows.append(row)
        
        # Создаем DataFrame
        df = pd.DataFrame(rows)
        
        # Сохраняем в Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Автонастройка ширины колонок
        if self.export_settings['auto_adjust_columns']:
            self._auto_adjust_columns(writer, sheet_name, df)
        
        self.logger.info(f"✅ Лист '{sheet_name}' создан: {len(rows)} товаров, {len(df.columns)} колонок")
    
    def _collect_all_characteristics(self, products: List[Dict]) -> List[str]:
        """
        Сбор всех уникальных характеристик из всех товаров
        
        Returns:
            List[str]: Отсортированный список названий характеристик
        """
        all_chars = set()
        
        for product in products:
            characteristics = product.get('characteristics', {})
            all_chars.update(characteristics.keys())
        
        return sorted(list(all_chars))
    
    def _format_images_field(self, processed_images: Dict[str, Any]) -> str:
        """
        Форматирование поля с фотографиями для Excel
        
        Args:
            processed_images: Данные обработанных изображений
            
        Returns:
            str: Строка с именами файлов через точку с запятой
        """
        if not processed_images:
            return ''
        
        main_image = processed_images.get('main_image', {})
        gallery_images = processed_images.get('gallery_images', [])
        
        all_images = []
        if main_image:
            all_images.append(main_image.get('file_name', ''))
        
        for img in gallery_images:
            all_images.append(img.get('file_name', ''))
        
        return '; '.join(all_images) if all_images else ''
    
    def _add_products_sheet(self, writer: pd.ExcelWriter, products: List[Dict], sheet_name: str):
        """Старый метод для обратной совместимости"""
        if not products:
            return
        
        # Форматирование данных товаров
        formatted_products = self.data_formatter.format_products_for_export(products)
        df = pd.DataFrame(formatted_products)
        
        # Сохранение в Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Автонастройка ширины колонок
        if self.export_settings['auto_adjust_columns']:
            self._auto_adjust_columns(writer, sheet_name, df)
    
    def _add_price_comparison_sheet(self, writer: pd.ExcelWriter, price_comparison: Dict, sheet_name: str):
        """Добавление листа сравнения цен"""
        comparison_data = []
        
        # Измененные товары
        for changed_product in price_comparison.get('changed_products', []):
            price_change = changed_product.get('price_change', {})
            comparison_data.append({
                'Товар': changed_product.get('name', ''),
                'Артикул': changed_product.get('article', ''),  # НОВОЕ ПОЛЕ
                'Старая цена': price_change.get('old_price'),
                'Новая цена': price_change.get('new_price'),
                'Изменение (%)': price_change.get('change_percent', 0),
                'Изменение (руб)': price_change.get('change_amount', 0),
                'Тип изменения': 'Увеличение' if price_change.get('change_direction') == 'increase' else 'Снижение',
                'Значительное': 'Да' if price_change.get('significant') else 'Нет'
            })
        
        # Новые товары
        for new_product in price_comparison.get('new_products', []):
            comparison_data.append({
                'Товар': new_product.get('name', ''),
                'Артикул': new_product.get('article', ''),  # НОВОЕ ПОЛЕ
                'Старая цена': 'Н/Д',
                'Новая цена': new_product.get('price'),
                'Изменение (%)': 'Новый товар',
                'Изменение (руб)': 'Новый товар',
                'Тип изменения': 'Новый товар',
                'Значительное': 'Н/Д'
            })
        
        if comparison_data:
            df = pd.DataFrame(comparison_data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._auto_adjust_columns(writer, sheet_name, df)
    
    def _add_trend_analysis_sheet(self, writer: pd.ExcelWriter, trend_analysis: Dict, sheet_name: str):
        """Добавление листа анализа трендов"""
        trends_data = []
        
        for product_key, analysis in trend_analysis.get('trends_analysis', {}).items():
            trends_data.append({
                'Товар': analysis.get('product_key', ''),
                'Артикул': analysis.get('article', ''),  # НОВОЕ ПОЛЕ
                'Тренд': self._translate_trend(analysis.get('trend_direction', '')),
                'Изменение (%)': analysis.get('total_change_percent', 0),
                'Текущая цена': analysis.get('price_range', {}).get('current', 0),
                'Минимальная цена': analysis.get('price_range', {}).get('min', 0),
                'Максимальная цена': analysis.get('price_range', {}).get('max', 0),
                'Волатильность': analysis.get('volatility', 0),
                'Точек данных': analysis.get('data_points', {}).get('total', 0),
                'Период (дни)': analysis.get('data_points', {}).get('period_days', 0)
            })
        
        if trends_data:
            df = pd.DataFrame(trends_data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._auto_adjust_columns(writer, sheet_name, df)
    
    def _add_validation_sheet(self, writer: pd.ExcelWriter, validation_results: Dict, sheet_name: str):
        """Добавление листа валидации"""
        validation_data = []
        
        # Валидные товары
        for valid_product in validation_results.get('valid_products', []):
            product = valid_product.get('product', {})
            validation_data.append({
                'Товар': product.get('name', ''),
                'Артикул': product.get('article', ''),  # НОВОЕ ПОЛЕ
                'Статус': '✅ Валиден',
                'Ошибки': 'Нет',
                'Предупреждения': '; '.join(valid_product.get('warnings', [])),
                'Цена': product.get('price', ''),
                'URL': product.get('url', '')
            })
        
        # Невалидные товары
        for invalid_product in validation_results.get('invalid_products', []):
            product = invalid_product.get('product', {})
            validation_data.append({
                'Товар': product.get('name', ''),
                'Артикул': product.get('article', ''),  # НОВОЕ ПОЛЕ
                'Статус': '❌ Невалиден',
                'Ошибки': '; '.join(invalid_product.get('errors', [])),
                'Предупреждения': '; '.join(invalid_product.get('warnings', [])),
                'Цена': product.get('price', ''),
                'URL': product.get('url', '')
            })
        
        if validation_data:
            df = pd.DataFrame(validation_data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._auto_adjust_columns(writer, sheet_name, df)
    
    def _add_alerts_sheet(self, writer: pd.ExcelWriter, alerts: Dict, sheet_name: str):
        """Добавление листа уведомлений"""
        alerts_data = []
        
        alert_types = {
            'price_alerts': '💰 Цены',
            'availability_alerts': '📦 Наличие',
            'trend_alerts': '📈 Тренды',
            'quality_alerts': '🔍 Качество данных'
        }
        
        for alert_type, alert_list in alerts.items():
            if isinstance(alert_list, list):
                for alert in alert_list:
                    alerts_data.append({
                        'Тип': alert_types.get(alert_type, alert_type),
                        'Уровень': alert.get('level', 'UNKNOWN'),
                        'Товар': alert.get('product', 'Общий'),
                        'Артикул': alert.get('article', ''),  # НОВОЕ ПОЛЕ
                        'Сообщение': alert.get('message', ''),
                        'Детали': str(alert.get('details', {}))
                    })
        
        if alerts_data:
            df = pd.DataFrame(alerts_data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            self._auto_adjust_columns(writer, sheet_name, df)
    
    def _add_summary_sheet(self, writer: pd.ExcelWriter, products: List[Dict], 
                          price_comparison: Optional[Dict], trend_analysis: Optional[Dict],
                          validation_results: Optional[Dict], alerts: Optional[Dict],
                          site_url: str):
        """Добавление сводного листа"""
        summary_data = []
        
        # Основная статистика
        summary_data.append({'Параметр': 'Общая информация', 'Значение': ''})
        summary_data.append({'Параметр': 'Дата отчета', 'Значение': datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        summary_data.append({'Параметр': 'URL сайта', 'Значение': site_url})
        summary_data.append({'Параметр': 'Всего товаров', 'Значение': len(products)})
        
        # НОВАЯ СТАТИСТИКА: Артикулы и изображения
        articles_generated = len([p for p in products if p.get('article')])
        images_processed = sum([len(p.get('processed_images', {}).get('processed_images', [])) for p in products])
        
        summary_data.append({'Параметр': 'Сгенерировано артикулов', 'Значение': articles_generated})
        summary_data.append({'Параметр': 'Обработано изображений', 'Значение': images_processed})
        
        # Статистика цен
        if price_comparison and price_comparison.get('success'):
            price_changes = price_comparison.get('price_changes', {})
            summary_data.append({'Параметр': '', 'Значение': ''})
            summary_data.append({'Параметр': 'Анализ цен', 'Значение': ''})
            summary_data.append({'Параметр': 'Товаров с изменением цены', 
                               'Значение': price_changes.get('increased', 0) + price_changes.get('decreased', 0)})
            summary_data.append({'Параметр': 'Увеличилось в цене', 'Значение': price_changes.get('increased', 0)})
            summary_data.append({'Параметр': 'Снизилось в цене', 'Значение': price_changes.get('decreased', 0)})
            summary_data.append({'Параметр': 'Новых товаров', 
                               'Значение': len(price_comparison.get('new_products', []))})
        
        # Качество данных
        if validation_results and validation_results.get('success'):
            metrics = validation_results.get('quality_metrics', {})
            summary_data.append({'Параметр': '', 'Значение': ''})
            summary_data.append({'Параметр': 'Качество данных', 'Значение': ''})
            summary_data.append({'Параметр': 'Оценка качества', 
                               'Значение': f"{metrics.get('quality_score', 0):.2%}"})
            summary_data.append({'Параметр': 'Валидных товаров', 
                               'Значение': metrics.get('valid_products_count', 0)})
            summary_data.append({'Параметр': 'Невалидных товаров', 
                               'Значение': metrics.get('invalid_products_count', 0)})
        
        # Уведомления
        if alerts:
            summary = alerts.get('summary', {})
            summary_data.append({'Параметр': '', 'Значение': ''})
            summary_data.append({'Параметр': 'Уведомления', 'Значение': ''})
            summary_data.append({'Параметр': 'Всего уведомлений', 'Значение': summary.get('total_alerts', 0)})
            summary_data.append({'Параметр': 'Критических', 'Значение': summary.get('critical_alerts', 0)})
            summary_data.append({'Параметр': 'Предупреждений', 'Значение': summary.get('warning_alerts', 0)})
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Сводка', index=False)
        self._auto_adjust_columns(writer, 'Сводка', df)
    
    def _translate_trend(self, trend: str) -> str:
        """Перевод тренда на русский"""
        trends = {
            'upward': '📈 Восходящий',
            'downward': '📉 Нисходящий', 
            'stable': '➡️ Стабильный',
            'uncertain': '❓ Неопределенный'
        }
        return trends.get(trend, trend)
    
    def _auto_adjust_columns(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
        """Автонастройка ширины колонок"""
        try:
            worksheet = writer.sheets[sheet_name]
            
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).str.len().max(),
                    len(str(col))
                )
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
                
        except Exception as e:
            self.logger.warning(f"⚠️ Не удалось настроить колонки: {e}")
    
    def _apply_advanced_formatting(self, filepath: str):
        """Применение продвинутого форматирования"""
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Форматирование заголовков
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Форматирование данных
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
            
            workbook.save(filepath)
            self.logger.info("🎨 Применено продвинутое форматирование")
            
        except Exception as e:
            self.logger.warning(f"⚠️ Не удалось применить форматирование: {e}")
    
    def generate_quick_export(self, products: List[Dict], filename: str = None) -> str:
        """
        Быстрый экспорт только товаров
        
        Args:
            products: Список товаров
            filename: Имя файла (опционально)
            
        Returns:
            str: Путь к файлу
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"quick_export_{timestamp}.xlsx"
        
        filepath = Path("data/output/excel_exports") / filename
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # ИСПОЛЬЗУЕМ НОВУЮ СТРУКТУРУ ДЛЯ БЫСТРОГО ЭКСПОРТА
        if self.export_settings['new_structure']:
            all_characteristics = self._collect_all_characteristics(products)
            rows = []
            for product in products:
                row = {
                    'Категория': product.get('category', ''),
                    'Артикул': product.get('article', ''),
                    'Фотографии': self._format_images_field(product.get('processed_images', {})),
                    'Описание': product.get('description', ''),
                    'Цена': product.get('price', '')
                }
                
                characteristics = product.get('characteristics', {})
                for char_name in all_characteristics:
                    row[char_name] = characteristics.get(char_name, '')
                
                rows.append(row)
            
            df = pd.DataFrame(rows)
        else:
            df = pd.DataFrame(products)
        
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        self.logger.info(f"💾 Быстрый экспорт создан: {filepath}")
        return str(filepath)
